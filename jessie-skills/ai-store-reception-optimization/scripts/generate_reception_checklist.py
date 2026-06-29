#!/usr/bin/env python3
"""
店铺承接优化动作清单生成器
根据SOP流程生成完整的Excel工作簿，包含所有输出交付物。
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def apply_header_style(ws, row, max_col, fill_color="4472C4"):
    """应用表头样式"""
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def apply_data_style(ws, start_row, end_row, max_col):
    """应用数据区域样式"""
    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border


def add_title(ws, title, row=1, col=1, max_col=7):
    """添加标题行"""
    title_font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    title_alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=col)
    cell.value = title
    cell.font = title_font
    cell.alignment = title_alignment


def add_subtitle(ws, text, row=2, col=1, max_col=7):
    """添加副标题行"""
    sub_font = Font(name="微软雅黑", size=9, color="808080")
    sub_alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = sub_font
    cell.alignment = sub_alignment


def create_action_checklist_sheet(wb):
    """创建：店铺承接优化动作清单"""
    ws = wb.create_sheet("动作清单")
    add_title(ws, "店铺承接优化动作清单")
    add_subtitle(ws, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")

    headers = ["优化模块", "当前问题", "具体优化动作", "优先级", "是否618前必做", "是否适合60天持续优化", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4, 7)

    modules = ["标题与首屏", "商品页卖点承接", "店铺首页承接", "活动页与直播间承接", "利益点与转化动作"]
    for i, mod in enumerate(modules):
        ws.cell(row=5 + i, column=1, value=mod)

    apply_data_style(ws, 5, 9, 7)
    col_widths = [22, 35, 40, 12, 18, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def create_priority_actions_sheet(wb):
    """创建：优先动作表"""
    ws = wb.create_sheet("优先动作")
    add_title(ws, "Top3立即优化 / 618前必做 / 60天持续优化", max_col=6)
    add_subtitle(ws, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", max_col=6)

    headers = ["类型", "动作名称", "为什么是它", "先怎么推进", "责任团队", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4, 6)

    types = ["Top1立即优化", "Top2立即优化", "Top3立即优化", "618前必做动作", "60天持续优化动作"]
    for i, t in enumerate(types):
        ws.cell(row=5 + i, column=1, value=t)

    apply_data_style(ws, 5, 9, 6)
    col_widths = [20, 30, 35, 35, 20, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def create_team_actions_sheet(wb):
    """创建：团队动作分工表"""
    ws = wb.create_sheet("团队动作分工")
    add_title(ws, "团队动作分工表", max_col=5)
    add_subtitle(ws, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", max_col=5)

    headers = ["团队", "当前最该先做什么", "依赖谁配合", "负责人/时间", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4, 5)

    teams = ["内容团队", "电商团队", "直播团队", "运营团队"]
    for i, t in enumerate(teams):
        ws.cell(row=5 + i, column=1, value=t)

    apply_data_style(ws, 5, 8, 5)
    col_widths = [16, 35, 25, 25, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def create_boss_summary_sheet(wb):
    """创建：老板版摘要"""
    ws = wb.create_sheet("老板版摘要")
    add_title(ws, "老板版摘要", max_col=2)

    items = [
        ("一句话结论", "承接不是一句空话，而是一组必须被拆出来、立刻推进的动作。"),
        ("当前最需要立即优化的3项承接动作", ""),
        ("必须在618前立刻完成的动作", ""),
        ("适合放进60天训练营持续优化的动作", ""),
        ("为什么这几个动作最关键", ""),
        ("本周最该先推进的动作", ""),
    ]

    label_font = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
    label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    data_font = Font(name="微软雅黑", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for i, (label, value) in enumerate(items):
        row = 3 + i
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = label_font
        cell_a.fill = label_fill
        cell_a.border = thin_border
        cell_a.alignment = Alignment(vertical="center")

        cell_b = ws.cell(row=row, column=2, value=value)
        cell_b.font = data_font
        cell_b.border = thin_border
        cell_b.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 60
    return ws


def create_sop_main_sheet(wb):
    """创建：SOP主表（20步流程）"""
    ws = wb.create_sheet("SOP主表")
    add_title(ws, "AI工具执行：店铺承接优化动作清单 SOP", max_col=7)
    add_subtitle(ws, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", max_col=7)

    headers = ["阶段", "步骤", "任务目标", "具体动作", "输入内容示例", "输出结果", "注意事项"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4, 7)

    steps = [
        (0, "明确本次任务", "统一认知：承接不是只改商品页，而是优化整套店铺承接系统",
         "明确目标：围绕核心SKU，列出一组可以立刻推进的店铺承接优化动作",
         "品牌名、核心SKU、课程主题、当前店铺承接现状",
         "店铺承接优化任务定义",
         "开场要强调：承接不是一句空话，而是要变成一组明确动作"),
        (1, "回看当前店铺承接现状", "先看目前店铺到底是怎么接用户的",
         "汇总当前商品页、店铺首页、活动页、直播间、导购内容、利益点结构的现状",
         "商品页链接、店铺首页、活动页、直播间、店铺装修图",
         "当前承接现状表",
         "不先看清现状，动作就会很空"),
        (2, "锁定核心SKU", "明确这次先围绕哪个产品来优化承接",
         "选择一个最重要的核心SKU作为承接优化主对象",
         "SKU清单、主推款、流量重点、利润重点",
         "核心SKU确认表",
         "先优化最关键的一条承接路径，不要一开始全部铺开"),
        (3, "回收前面内容路径", "先把内容里种下的东西拿回来对照",
         "汇总内容里已经种下的搜索词、卖点词、场景、活动预期、下一步动作",
         "内容样本、外种内收链路图、内容到承接检查表",
         "内容输入信息表",
         "承接优化必须接着前面的内容逻辑来做"),
        (4, "检查标题与首屏是否接住", "判断标题和首屏能不能接住内容种下的心智",
         "让AI检查标题是否能接住搜索词和卖点词，首屏是否继续放大购买理由",
         "商品标题、首屏截图、关键词、卖点、内容表达",
         "标题与首屏检查表",
         "用户进来第一眼接不住，后面很难继续推进"),
        (5, "梳理标题与首屏优化动作", "把问题变成可执行动作",
         "让AI输出标题改法、首屏信息优化、首屏利益点强化建议",
         "当前标题、首屏内容、搜索词、内容关键词",
         "标题与首屏优化动作表",
         "这类动作通常最容易先做，也最容易快速见效"),
        (6, "检查商品页卖点承接是否清楚", "判断商品页是不是在延续内容逻辑",
         "检查商品页卖点层级是否清楚、是否与外部内容保持一致",
         "商品详情页、卖点表、内容表达、竞品页面",
         "商品页卖点检查表",
         "如果商品页和内容讲的不是一套逻辑，用户会明显掉线"),
        (7, "梳理商品页卖点优化动作", "把商品页问题变成可落地动作",
         "让AI输出商品页卖点重排、层级重构、表达统一建议",
         "商品详情页、卖点结构、外部内容逻辑",
         "商品页卖点优化动作表",
         '重点不是"多写"，而是"更顺、更统一"'),
        (8, "检查店铺首页承接是否延续内容逻辑", "看店铺首页是不是在继续接，而不是只是堆商品",
         "让AI检查店铺首页是否围绕内容逻辑、核心SKU和主场景延续表达",
         "店铺首页、店铺装修、店铺推荐位、活动入口",
         "店铺首页承接检查表",
         "很多店铺首页的问题不是没有东西，而是没有逻辑"),
        (9, "梳理店铺首页优化动作", '让首页承接从"堆商品"变成"接逻辑"',
         "让AI输出首页模块调整、主推区优化、氛围统一、场景延续建议",
         "店铺首页截图、模块结构、主推商品、活动入口",
         "店铺首页优化动作表",
         "店铺首页的任务不是展示全部，而是继续接住前面内容逻辑"),
        (10, "检查活动页与直播间承接是否顺", "看活动页和直播间能不能让用户快速找到下一步动作",
         "让AI检查活动页、直播间入口、直播话术、利益点表达是否和内容目标一致",
         "活动页、直播间脚本、直播入口、内容导向",
         "活动页与直播间检查表",
         "用户被内容打动后，必须能快速进入下一步动作"),
        (11, "梳理活动页与直播间优化动作", "把活动页和直播间问题转成动作",
         "让AI输出活动页承接优化、直播间入口优化、直播利益点强化建议",
         "活动页、直播脚本、直播节奏、内容路径",
         "活动页与直播间优化动作表",
         '这一步是从"被种草"走向"被收割"的关键节点'),
        (12, "检查利益点与转化动作是否清晰", "判断优惠、赠品、机制、按钮、引导路径是否足够明确",
         "让AI检查当前页面的利益点和动作引导是否让用户清楚知道下一步做什么",
         "优惠信息、赠品机制、按钮文案、转化路径",
         "利益点与动作检查表",
         "页面有信息不等于用户真的知道该怎么行动"),
        (13, "梳理利益点与转化优化动作", "把模糊利益点变成清晰动作",
         "让AI输出优惠表达优化、按钮文案优化、转化路径优化建议",
         "页面利益点、CTA、路径图、活动机制",
         "利益点优化动作表",
         '转化动作一定要清晰到"用户下一步要做什么"'),
        (14, "汇总全部承接优化动作", "把分散问题整理成行动清单",
         "将标题首屏、商品页、店铺首页、活动页、直播间、利益点动作汇总成一张清单",
         "前面所有检查和动作建议",
         "店铺承接优化动作清单初版",
         "这是第三张现场输出物的核心表"),
        (15, "判断最需要立即优化的3项动作", "帮团队先聚焦真正重要的动作",
         "让AI帮助筛出当前最需要立即优化的3项店铺承接动作",
         "动作清单、618目标、资源限制、页面现状",
         "Top3立即优化动作表",
         '一定要逼团队聚焦，不要列一堆"都重要"'),
        (16, "判断必须在618前立刻完成的动作", "找出最不能拖的动作",
         "让AI判断哪一项必须在618前立刻完成，否则会直接影响结果",
         "Top3动作、节点时间、活动节奏、转化链路",
         "618前必做动作表",
         "这是短期必须落地的承接动作"),
        (17, "判断哪一项适合放进60天训练营持续优化", "区分短期动作和长期动作",
         "让AI判断哪一项更适合作为训练营中的持续优化项目",
         "全部动作清单、资源、训练营周期、执行难度",
         "60天持续优化动作表",
         "不是所有动作都必须今晚做完，有些适合持续优化"),
        (18, "分配到团队动作", "让动作清单真正进入执行",
         "拆成内容团队、电商团队、直播团队、运营团队各自先做什么",
         "动作清单、岗位分工、页面责任",
         "团队动作分工表",
         "不拆到团队，动作清单就很难真正推进"),
        (19, "建立后续复盘接口", "后续验证承接优化有没有生效",
         "定义后续要看的信号：搜索回流、商品页停留、店铺页访问、直播间点击、加购率、下单率等",
         "指标口径、复盘节奏、阶段目标",
         "承接优化复盘指标表",
         "承接优化一定要进入后续复盘，才能越改越准"),
    ]

    for i, (phase, step, goal, action, inp, out, note) in enumerate(steps):
        row = 5 + i
        ws.cell(row=row, column=1, value=phase)
        ws.cell(row=row, column=2, value=step)
        ws.cell(row=row, column=3, value=goal)
        ws.cell(row=row, column=4, value=action)
        ws.cell(row=row, column=5, value=inp)
        ws.cell(row=row, column=6, value=out)
        ws.cell(row=row, column=7, value=note)

    apply_data_style(ws, 5, 24, 7)
    col_widths = [8, 36, 44, 48, 38, 30, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def create_knowledge_sheet(wb):
    """创建：标准理解与关键认知"""
    ws = wb.create_sheet("标准理解与认知")

    # 标题
    add_title(ws, "店铺承接优化 — 标准理解与关键认知", max_col=3)

    # 标准理解
    ws.merge_cells("A4:C4")
    ws.cell(row=4, column=1, value="标准理解").font = Font(name="微软雅黑", bold=True, size=12, color="1F4E79")

    headers = ["模块", "它回答的问题", "课堂解释"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    apply_header_style(ws, 5, 3)

    standards = [
        ("标题与首屏优化", "用户进来第一眼有没有被接住", "标题要接住搜索词和卖点词，首屏要继续放大购买理由"),
        ("商品页卖点承接", "商品页是不是在继续说内容里的那套逻辑", "卖点表达要一致，层级要清楚"),
        ("店铺首页承接", "店铺首页是不是在继续延续内容逻辑", "不能只是无序商品堆砌，而要接住主场景和主逻辑"),
        ("活动页与直播间承接", "用户被打动后有没有快速找到下一步动作", "活动页、直播间要继续承接内容引导"),
        ("利益点与转化动作", "用户有没有明确理由和路径继续往前走", "优惠、赠品、机制、按钮、导流路径都必须足够清晰"),
    ]
    for i, (mod, q, exp) in enumerate(standards):
        row = 6 + i
        ws.cell(row=row, column=1, value=mod)
        ws.cell(row=row, column=2, value=q)
        ws.cell(row=row, column=3, value=exp)
    apply_data_style(ws, 6, 10, 3)

    # 关键认知
    ws.merge_cells("A12:C12")
    ws.cell(row=12, column=1, value="关键认知（3个）").font = Font(name="微软雅黑", bold=True, size=12, color="1F4E79")

    headers2 = ["#", "关键认知"]
    for col, h in enumerate(headers2, 1):
        ws.cell(row=13, column=col, value=h)
    apply_header_style(ws, 13, 2)

    knowledges = [
        "承接不是只改商品页 — 真正接住用户的是整套店铺承接系统，不是单页优化。",
        "店铺承接优化不是一句空话，而是一组动作 — 要把承接问题拆成标题、首屏、商品页、首页、活动页、直播间、利益点等具体动作。",
        '真正的承接优化，不是"多"，而是"顺" — 用户被内容打动后，能不能顺着路径继续往前走，才是关键。',
    ]
    for i, k in enumerate(knowledges):
        row = 14 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=k)
    apply_data_style(ws, 14, 16, 2)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 55
    return ws


def create_deliverables_sheet(wb):
    """创建：课堂交付物清单"""
    ws = wb.create_sheet("交付物清单")
    add_title(ws, "商家做完后至少要沉淀出的交付物", max_col=2)

    deliverables = [
        ("当前承接现状表", "先看现在店铺是怎么接用户的"),
        ("核心SKU确认表", "明确承接优化的核心对象"),
        ("内容输入信息表", "回收前面内容种下的搜索词和卖点词"),
        ("标题与首屏检查表", "判断标题和首屏是否接住内容心智"),
        ("标题与首屏优化动作表", "明确最先可做的首屏优化动作"),
        ("商品页卖点检查表", "判断商品页是否在延续内容卖点"),
        ("商品页卖点优化动作表", "明确商品页卖点层级和表达优化动作"),
        ("店铺首页承接检查表", "判断首页是否延续内容逻辑"),
        ("店铺首页优化动作表", "明确首页模块和氛围优化动作"),
        ("活动页与直播间检查表", "判断活动页和直播间是否承接顺畅"),
        ("活动页与直播间优化动作表", "明确活动页和直播间优化动作"),
        ("利益点与动作检查表", "判断优惠、赠品、CTA是否清晰"),
        ("利益点优化动作表", "明确利益点优化方向"),
        ("店铺承接优化动作清单", "★★★ 第三张现场输出物的核心表"),
        ("Top3立即优化动作表", "聚焦当前最重要的3项动作"),
        ("618前必做动作表", "短期必须落地的承接动作"),
        ("60天持续优化动作表", "适合长期持续优化的项目"),
        ("团队动作分工表", "让动作真正进入执行"),
        ("承接优化复盘指标表", "定义后续验证指标"),
        ("老板版摘要", "给老板看的一页纸总结"),
    ]

    headers = ["交付物名称", "说明"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    apply_header_style(ws, 3, 2)

    fill_light = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for i, (name, desc) in enumerate(deliverables):
        row = 4 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=desc)
        if "★★★" in desc:
            ws.cell(row=row, column=1).fill = fill_light
            ws.cell(row=row, column=2).fill = fill_light

    apply_data_style(ws, 4, 23, 2)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 50
    return ws


def main(output_dir=None):
    """主函数：生成店铺承接优化动作清单Excel"""
    if output_dir is None:
        output_dir = os.getcwd()

    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 按顺序创建所有sheet
    create_sop_main_sheet(wb)
    create_knowledge_sheet(wb)
    create_action_checklist_sheet(wb)
    create_priority_actions_sheet(wb)
    create_team_actions_sheet(wb)
    create_boss_summary_sheet(wb)
    create_deliverables_sheet(wb)

    output_path = os.path.join(output_dir, "店铺承接优化动作清单.xlsx")
    wb.save(output_path)
    print(f"[OK] 店铺承接优化动作清单已生成：{output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else None
    main(out)
