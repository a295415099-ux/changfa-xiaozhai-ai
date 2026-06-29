#!/usr/bin/env python3
"""
AI信任证据链建设 - 批量输出生成脚本
基于步骤1-7的盘点结果，批量生成步骤8-14的输出表格（Excel格式）。
"""

import argparse
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_sheet(ws, headers, col_widths):
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_rows(ws, data, start_row=2):
    for r_idx, row in enumerate(data, start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


def create_workbook(data, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 证据链完整度检查表（步骤8）
    ws1 = wb.create_sheet('证据链完整度检查表')
    headers1 = ['核心顾虑', '评价证据(✅/❌)', '内容证据(✅/❌)', '品牌证据(✅/❌)', '服务证据(✅/❌)', '完整度评级']
    style_sheet(ws1, headers1, [25, 12, 12, 12, 12, 12])
    if data.get('integrity_check'):
        add_rows(ws1, data['integrity_check'])

    # Sheet 2: AI可引用表达表（步骤9）
    ws2 = wb.create_sheet('AI可引用表达表')
    headers2 = ['引用编号', '对应顾虑', '可引用表达', '引用场景', '引用强度']
    style_sheet(ws2, headers2, [10, 20, 40, 20, 10])
    if data.get('ai_expressions'):
        add_rows(ws2, data['ai_expressions'])

    # Sheet 3: 用户顾虑—证据—话术对应表（步骤10）
    ws3 = wb.create_sheet('顾虑—证据—话术对应表')
    headers3 = ['用户顾虑', '对应证据', '推荐话术', '适用角色', '适用渠道']
    style_sheet(ws3, headers3, [20, 25, 35, 15, 15])
    if data.get('concern_evidence_script'):
        add_rows(ws3, data['concern_evidence_script'])

    # Sheet 4: AI可引用品牌信任语料（步骤11）
    ws4 = wb.create_sheet('AI可引用品牌信任语料')
    headers4 = ['语料编号', '语料内容', '证据来源', '引用权重', '适用场景']
    style_sheet(ws4, headers4, [10, 40, 20, 10, 20])
    if data.get('trust_corpus'):
        add_rows(ws4, data['trust_corpus'])

    # Sheet 5: 证据真实性检查表（步骤12）
    ws5 = wb.create_sheet('证据真实性检查表')
    headers5 = ['证据条目', '可验证性', '是否夸大', '是否过时', '处理建议']
    style_sheet(ws5, headers5, [30, 12, 12, 12, 15])
    if data.get('authenticity_check'):
        add_rows(ws5, data['authenticity_check'])

    # Sheet 6: 推荐可信度检查表（步骤13）
    ws6 = wb.create_sheet('推荐可信度检查表')
    headers6 = ['检查维度', '当前状态', '是否满足', '缺失说明']
    style_sheet(ws6, headers6, [22, 30, 10, 25])
    if data.get('credibility_check'):
        add_rows(ws6, data['credibility_check'])

    # Sheet 7: AI信任证据链建设表（步骤14）
    ws7 = wb.create_sheet('AI信任证据链建设表（初版）')
    headers7 = ['核心购买顾虑', '对应信任证据', '评价证据', '内容证据', '品牌证据', '服务证据', 'AI可引用表达', '备注']
    style_sheet(ws7, headers7, [18, 15, 22, 22, 18, 18, 30, 15])
    if data.get('trust_chain'):
        add_rows(ws7, data['trust_chain'])

    # Sheet 8: 当前信任缺口清单（步骤15）
    ws8 = wb.create_sheet('当前信任缺口清单')
    headers8 = ['缺口类型', '影响范围', '紧急程度', '补齐建议', '数据来源']
    style_sheet(ws8, headers8, [15, 20, 10, 30, 20])
    if data.get('gaps'):
        add_rows(ws8, data['gaps'])

    # Sheet 9: 信任资产使用分工表（步骤16）
    ws9 = wb.create_sheet('信任资产使用分工表')
    headers9 = ['证据/话术类型', '使用团队', '使用场景', '使用方式', '更新频率']
    style_sheet(ws9, headers9, [18, 15, 20, 25, 12])
    if data.get('division'):
        add_rows(ws9, data['division'])

    # Sheet 10: 信任证据链更新机制表（步骤17）
    ws10 = wb.create_sheet('信任证据链更新机制表')
    headers10 = ['更新触发条件', '更新内容', '更新频率', '责任人', '关联步骤']
    style_sheet(ws10, headers10, [18, 25, 12, 15, 12])
    if data.get('update_mechanism'):
        add_rows(ws10, data['update_mechanism'])

    wb.save(output_path)
    print(json.dumps({'status': 'success', 'output': output_path, 'sheets': len(wb.sheetnames)}))


def main():
    parser = argparse.ArgumentParser(description='AI信任证据链建设 - 批量输出生成')
    parser.add_argument('--input', '-i', help='JSON输入文件路径，包含各步骤数据')
    parser.add_argument('--output', '-o', required=True, help='输出Excel文件路径')
    args = parser.parse_args()

    data = {}
    if args.input:
        with open(args.input, 'r', encoding='utf-8') as f:
            data = json.load(f)

    create_workbook(data, args.output)


if __name__ == '__main__':
    main()
