#!/usr/bin/env python3
"""生成AI推荐内收链路检查表（Excel格式）"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_check_table(output_path, brand_name="待填写"):
    wb = Workbook()

    # Sheet 1: AI推荐内收链路检查表
    ws1 = wb.active
    ws1.title = "AI推荐内收链路检查表"
    _setup_check_table(ws1, brand_name)

    # Sheet 2: 外种-AI问答-进店承接优化清单
    ws2 = wb.create_sheet("优化清单")
    _setup_optimization_list(ws2)

    # Sheet 3: 7天链路修复行动表
    ws3 = wb.create_sheet("7天修复行动")
    _setup_7day_action(ws3)

    # Sheet 4: 链路断点优先级表
    ws4 = wb.create_sheet("断点优先级")
    _setup_priority_table(ws4)

    # Sheet 5: 链路执行分工表
    ws5 = wb.create_sheet("执行分工")
    _setup_execution_table(ws5)

    wb.save(output_path)
    print(f"✅ 检查表已生成: {output_path}")
    return output_path

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
wrap_align = Alignment(vertical="top", wrap_text=True)

def _style_header(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = center_align
        cell.border = thin_border

def _style_cell(ws, row, col, align=None):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = align or wrap_align

def _setup_check_table(ws, brand_name):
    ws.merge_cells("A1:H1")
    ws["A1"] = f"AI推荐内收链路检查表 — {brand_name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["检查模块", "当前情况", "是否存在断点", "主要问题", "优化动作", "负责人", "优先级", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 8)

    modules = [
        "外部内容是否清楚",
        "用户问题是否覆盖",
        "商品信息是否一致",
        "推荐理由是否明确",
        "信任证据是否充足",
        "承接页面是否匹配",
        "客服是否能接住",
    ]
    for idx, mod in enumerate(modules):
        row = 3 + idx
        ws.cell(row=row, column=1, value=mod)
        for c in range(1, 9):
            _style_cell(ws, row, c)

    cols = [28, 20, 14, 25, 25, 10, 10, 15]
    for i, w in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _setup_optimization_list(ws):
    ws.merge_cells("A1:E1")
    ws["A1"] = "外种—AI问答—进店承接优化清单"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["优化项目", "当前问题", "优化动作", "预期效果", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 5)

    categories = [
        "外部内容优化", "用户问题覆盖", "商品信息一致性",
        "推荐理由强化", "信任证据补充", "承接页面优化", "客服承接优化"
    ]
    for idx, cat in enumerate(categories):
        row = 3 + idx
        ws.cell(row=row, column=1, value=cat)
        for c in range(1, 6):
            _style_cell(ws, row, c)

    for i, w in enumerate([25, 30, 30, 25, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _setup_7day_action(ws):
    ws.merge_cells("A1:E1")
    ws["A1"] = "7天链路修复行动表"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["天数", "修复动作", "谁负责", "何时完成", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 5)

    for day in range(1, 8):
        row = 2 + day
        ws.cell(row=row, column=1, value=f"第{day}天")
        for c in range(1, 6):
            _style_cell(ws, row, c)

    for i, w in enumerate([10, 35, 12, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _setup_priority_table(ws):
    ws.merge_cells("A1:F1")
    ws["A1"] = "链路断点优先级表"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["序号", "断点位置", "严重程度", "影响范围", "优先级", "修复方向"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 6)

    for idx in range(1, 6):
        row = 2 + idx
        ws.cell(row=row, column=1, value=idx)
        for c in range(1, 7):
            _style_cell(ws, row, c)

    for i, w in enumerate([8, 25, 12, 20, 10, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _setup_execution_table(ws):
    ws.merge_cells("A1:F1")
    ws["A1"] = "链路执行分工表"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["链路问题", "负责角色", "具体动作", "验收标准", "完成时间", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, 6)

    roles = ["内容团队", "电商团队", "客服团队", "投放团队", "品牌负责人"]
    for idx, role in enumerate(roles):
        row = 3 + idx
        ws.cell(row=row, column=2, value=role)
        for c in range(1, 7):
            _style_cell(ws, row, c)

    for i, w in enumerate([25, 12, 30, 25, 12, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "AI推荐内收链路检查表.xlsx"
    brand = sys.argv[2] if len(sys.argv) > 2 else "待填写"
    create_check_table(output, brand)
