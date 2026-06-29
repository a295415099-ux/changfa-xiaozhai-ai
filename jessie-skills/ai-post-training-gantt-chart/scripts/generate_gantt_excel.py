#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_FILE = os.path.join(os.path.dirname(__file__), 'gantt_data.json')

def style_header(ws, row, max_col, fill_color='4472C4', font_color='FFFFFF'):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    header_font = Font(name='Arial', bold=True, color=font_color, size=11)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def write_rows(ws, data, start_row=1):
    for r, row_data in enumerate(data, start_row):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val).alignment = Alignment(wrap_text=True)

def create_all_sheets(wb, data):
    # Sheet: 使用说明
    ws = wb.active
    ws.title = data['sheets'][0]['name']
    write_rows(ws, data['sheets'][0]['rows'])
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 80
    style_header(ws, 1, 2)

    # Sheet: 课程讲稿版
    ws = wb.create_sheet(data['sheets'][1]['name'])
    write_rows(ws, data['sheets'][1]['rows'])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 80

    # Sheet: SOP主表
    ws = wb.create_sheet(data['sheets'][2]['name'])
    headers = data['sheets'][2]['headers']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers), fill_color='2F5496')
    write_rows(ws, data['sheets'][2]['rows'], 2)
    for i, w in enumerate([8, 28, 40, 50, 35, 28, 45], 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Sheet: 标准理解
    ws = wb.create_sheet(data['sheets'][3]['name'])
    headers = data['sheets'][3]['headers']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 3)
    write_rows(ws, data['sheets'][3]['rows'], 2)
    for i, w in enumerate([18, 35, 50], 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Sheet: 关键认知
    ws = wb.create_sheet(data['sheets'][4]['name'])
    title = data['sheets'][4]['title']
    ws.merge_cells('A1:B1')
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    write_rows(ws, data['sheets'][4]['rows'], 3)
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 45

    # Sheet: 课堂交付物
    ws = wb.create_sheet(data['sheets'][5]['name'])
    title = data['sheets'][5]['title']
    ws.merge_cells('A1:B1')
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    write_rows(ws, data['sheets'][5]['rows'], 2)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40

    # Sheet: 甘特表模板
    ws = wb.create_sheet(data['sheets'][6]['name'])
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value=data['sheets'][6]['title']).font = Font(bold=True, size=12)
    headers = data['sheets'][6]['headers']
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    style_header(ws, 2, len(headers), fill_color='2F5496')
    for i in range(1, 11):
        ws.column_dimensions[chr(64+i)].width = 18
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 30

    # Sheet: 阶段与风险模板
    ws = wb.create_sheet(data['sheets'][7]['name'])
    headers = data['sheets'][7]['headers']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    write_rows(ws, data['sheets'][7]['rows'], 2)
    for i, w in enumerate([18, 30, 30, 18, 18, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Sheet: 滚动更新模板
    ws = wb.create_sheet(data['sheets'][8]['name'])
    headers = data['sheets'][8]['headers']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    write_rows(ws, data['sheets'][8]['rows'], 2)
    for i, w in enumerate([25, 35, 18, 22, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Sheet: 老板版摘要模板
    ws = wb.create_sheet(data['sheets'][9]['name'])
    headers = data['sheets'][9]['headers']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 2)
    write_rows(ws, data['sheets'][9]['rows'], 2)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 60

def main():
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    wb = Workbook()
    create_all_sheets(wb, data)
    output_path = sys.argv[1] if len(sys.argv) > 1 else '/tmp/ai_gantt_sop.xlsx'
    wb.save(output_path)
    print(f'Generated: {output_path}')

if __name__ == '__main__':
    main()
